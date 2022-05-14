from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import Notebook
import xlrd
from openpyxl import load_workbook
import pyperclip
from io import open
import os
import pickle

openFile=''
sheet=''
guardado=''
inventarioExcel= ''
hojaInventarioExcel=''
llavesExcel=''
hojaLlavesExcel=''
ingresoManual= ''

def cerrarVentanaParaIngresarValor(valorIngresado,ventanaSecundaria):
    global ingresoManual
    ingresoManual = valorIngresado
    ventanaSecundaria.destroy()

def ventanaParaIngresarValor(mensaje,anchoDeEntry=None):
    ventanaSecundaria = Toplevel()
    ventanaSecundaria.title('¡Información incompleta!')
    ventanaSecundaria.resizable(0,0)
    ventanaSecundaria.geometry('+813+457')
    ventanaSecundaria.transient(ventana)

    Label(ventanaSecundaria,text=mensaje).pack(pady=10,padx=30)
    entryDeAlerta = Entry(ventanaSecundaria,width=anchoDeEntry)
    entryDeAlerta.pack()
    Button(ventanaSecundaria,text='Aceptar',command=lambda: cerrarVentanaParaIngresarValor(entryDeAlerta.get(),ventanaSecundaria)).pack(pady=10,padx=10)
    ventanaSecundaria.grab_set()
    ventanaSecundaria.wait_window()

def copirait():
    messagebox.showinfo("About","Copyright © todos los derechos reservados a Julián Ferrari")

def verificarRutaXML(opcion,validacionConfig=False):
    if(os.path.os.path.isfile('Rutas.xml')) == True:
        archivo = open('Rutas.xml','rb')
        infoRutas = pickle.load(archivo)
        info = guardarCarpeta(opcion,infoRutas,validacionConfig)
        archivo.close()
        return info

    else:
        archivo = open ('Rutas.xml','wb')
        infoRutas = ['','','','','']
        pickle.dump(infoRutas,archivo)
        archivo.close()
        verificarRutaXML(opcion)                     

def guardarCarpeta(opcion,infoRutas,validacionConfig):
    global openFile,sheet,guardado,setCajaInfoRest,setCajaGuardar,setCajaPlantilla,excel_mod, inventarioExcel, hojaInventarioExcel,hojaLlavesExcel,llavesExcel
    if opcion == 1:
        if infoRutas[0] == '' or validacionConfig == True:
            guardado=filedialog.askopenfilename(title="Información de restaurantes",initialdir="C:/",filetypes=(("Archivos Excel","*.xls"),("Todos los archivos","*.*")))
            if(guardado == ''):
                return
            openFile = xlrd.open_workbook(guardado)
            sheet = openFile.sheet_by_name("Directorio-2020")
            infoRutas[0] = guardado
            archivo = open("Rutas.xml","wb")
            pickle.dump(infoRutas,archivo)
            archivo.close()
            setCajaInfoRest.set(guardado)
        else:
            openFile = xlrd.open_workbook(infoRutas[0])
            sheet = openFile.sheet_by_name("Directorio-2020")
            setCajaInfoRest.set(infoRutas[0])
        
    elif opcion == 2:
        if infoRutas[1] == '' or validacionConfig == True:
            guardado=filedialog.askopenfilename(title="Plantilla de layout",initialdir="C:/",filetypes=(("Archivos Excel","*.xlsx"),("Todos los archivos","*.*")))
            if(guardado == ''):
                return
            excel_mod = load_workbook(guardado)
            sheet = excel_mod['Hoja1']
            infoRutas[1] = guardado
            archivo = open("Rutas.xml","wb")
            pickle.dump(infoRutas,archivo)
            archivo.close()
            setCajaPlantilla.set(guardado)
        else:
            excel_mod = load_workbook(infoRutas[1])
            sheet = excel_mod['Hoja1']
            setCajaPlantilla.set(infoRutas[1])

    elif opcion == 3:
        if infoRutas[2] == '' or validacionConfig == True:
            guardado=filedialog.askdirectory(title="Guardar como layout")
            if(guardado == ''):
                return
            infoRutas[2] = guardado
            archivo=open("Rutas.xml","wb")
            pickle.dump(infoRutas,archivo)
            archivo.close()
            setCajaGuardar.set(guardado)
            return infoRutas[2]
        else:
            setCajaGuardar.set(infoRutas[2])
            return infoRutas[2]
    
    elif opcion == 4:
        if infoRutas[3] == '' or validacionConfig == True:
            guardado= filedialog.askopenfilename(title="Inventario de pinpad",initialdir="C:/",filetypes=(("Archivos Excel","*.xlsx"),("Todos los archivos","*.*")))
            if(guardado == ''):
                return
            inventarioExcel = load_workbook(guardado)
            hojaInventarioExcel = inventarioExcel['Inventario']
            infoRutas[3] = guardado
            archivo=open("Rutas.xml","wb")
            pickle.dump(infoRutas,archivo)
            archivo.close()
            setCajaInventario.set(guardado)
        else:
            inventarioExcel = load_workbook(infoRutas[3])
            hojaInventarioExcel = inventarioExcel['Inventario']
            setCajaInventario.set(infoRutas[3])

    elif opcion == 5:
        if infoRutas[4] == '' or validacionConfig == True:
            guardado=filedialog.askopenfilename(title="Llaves de pinpad",initialdir="C:/",filetypes=(("Archivos Excel","*.xls"),("Todos los archivos","*.*")))
            if(guardado == ''):
                return
            llavesExcel = xlrd.open_workbook(guardado)
            hojaLlavesExcel = llavesExcel.sheet_by_name("MCDON_ter")
            infoRutas[4] = guardado
            archivo=open("Rutas.xml","wb")
            pickle.dump(infoRutas,archivo)
            archivo.close()
            setCajaLlaves.set(guardado)
        else:
            llavesExcel = xlrd.open_workbook(infoRutas[4])
            hojaLlavesExcel = llavesExcel.sheet_by_name("MCDON_ter")
            setCajaLlaves.set(infoRutas[4])                          

def separaCadenaPorCaracter(cadena,separaValores,posicion):
    cadena = str(cadena)
    cadena = cadena.split(separaValores)
    return cadena[posicion]

def cargarReemplazo():#Cargar reemplazo de pinpad
    
    if altaReemplazo.get()!=1 and altaReemplazo.get()!= 2:
        messagebox.showerror("¡Error!","Seleccione tipo de solicitud")
        return

    global openFile,sheet, excel_mod, ingresoManual
    ingresoManual=''
    nombre_rest=''
    direccion = ''
    telefono = ''
    afiliado = ''
    verificarRutaXML(1)

    for i in range(sheet.nrows):
        excelAux = separaCadenaPorCaracter(sheet.cell_value(i,0),'.',0)
        if excelAux == cajaRestaurante.get().upper():
            nombre_rest = str(sheet.cell_value(i,1))
            try:
                afiliado=int(sheet.cell_value(i,8))
            except:
                ventanaParaIngresarValor("Favor de ingresar afiliación manualmente: ")
                afiliado = ingresoManual
            try:
                telefono=int(sheet.cell_value(i,10))
            except:
                ventanaParaIngresarValor("Favor de ingresar teléfono de contacto manualmente: ")
                telefono = ingresoManual
            if(str(sheet.cell_value(i,15))!=''):
                direccion = str(sheet.cell_value(i,15))
            else:
                ventanaParaIngresarValor("Favor de ingresar la dirección manualmente: ",'50')
                direccion = ingresoManual
            break

    if nombre_rest == '' :#Valida que haya ingresado un restaurante correcto
        messagebox.showerror("¡Atención!","Revisar que el restaurante ingresado sea correcto")
        return
    if altaReemplazo.get() == 2:
        if str(modelo.get()) != "Verifone VX820" and str(modelo.get()) != "Ingenico IPP320": #Valida si se selecciono el modelo de pinpad
            messagebox.showerror("¡Atención!","Selecione el modelo de pinpad")
            return

    verificarRutaXML(2)
    
    sheet['C5'] = nombre_rest
    sheet['C6'] = afiliado
    sheet['C11'] = telefono
    sheet['C8'] = direccion
    sheet['C17'] = cajaTicket.get()
    sheet['C16'] = cajaMotivo.get()

    if altaReemplazo.get() == 2:
        sheet['C15'] =  ("Marca: " + modelo.get() + " Numero de serie: " + cajaSerie.get())

    if altaReemplazo.get() == 2:
        if int(inventario.get()) !=1 and int(inventario.get()) !=2:
            messagebox.showerror("¡Atención!","Seleccione si se encuentra en inventario")
            return

    info = verificarRutaXML(3)
    
    excel_mod.save(str(info) + '/' + nombre_rest + " - " + cajaTicket.get() + " - Pos " + cajaPos.get() + " .xlsx")#Guardado de excel con nombre


    if altaReemplazo.get() == 2:
        if(int(inventario.get()))==2:
            pyperclip.copy(str(cajaTicket.get()) + """ - """ + nombre_rest + """ - Pinpad dañado - Pos """ + str(cajaPos.get()) + """
Adriana,

Nos contactamos desde la Mesa de Ayuda de Arcos Dorados para solicitar el reemplazo de un pinpad que se encuentra en inventario de EvoPayments
Falla: """ + str(cajaMotivo.get()) + """

Marca: """ + str(modelo.get()) + """
N/S: """ + str(cajaSerie.get()) + """

Adjunto el archivo LayOut con los detalles del restaurante y contacto para efectuar el reemplazo. 

Cualquier información adicional que necesiten no dude en contactarnos.

Saludos,""")
        
        if(int(inventario.get()))==1:
            pyperclip.copy(str(cajaTicket.get()) + """ - """ + nombre_rest + """ - Pinpad dañado - Pos """ + str(cajaPos.get()) + """
Adriana,

Nos contactamos desde la Mesa de Ayuda de Arcos Dorados para solicitar el alta de un pinpad. El restaurante posee uno que pertenece a Grupo Asesores ya que no se encuentra en el inventario de EvoPayments.
Falla: """ + str(cajaMotivo.get()) + """

Marca: """ + str(modelo.get()) + """
N/S: """ + str(cajaSerie.get()) + """

Adjunto el archivo LayOut con los detalles del restaurante y contacto para efectuar el reemplazo. 

Cualquier información adicional que necesiten no dude en contactarnos.

Saludos,""")
    elif altaReemplazo.get() == 1:
        pyperclip.copy(str(cajaTicket.get()) + """ - """ + ' ' + nombre_rest + """ - Alta de pinpad - Pos """ + str(cajaPos.get()) + """
 estimado,

Nos contactamos desde la Mesa de Ayuda de Arcos Dorados para solicitar el alta de un nuevo pinpad para el restaurante """ + nombre_rest + """, ya que la pos no cuenta con uno.

Adjunto el archivo LayOut con los detalles del restaurante y contacto para efectuar el Alta. 

Cualquier información adicional que necesiten no dude en contactarnos.

Saludos,
"""     )

def buscarPinpadInventario():
    numberSerial = str(cajaInventario.get())
    verificarRutaXML(1)
    verificarRutaXML(4)
    numAfiliacion = ''
    restaurante = ''
    for a in range(1,3000):
        if (str(hojaInventarioExcel[f'A{a}'].value) == numberSerial.upper()):
            numAfiliacion = str (hojaInventarioExcel[f'E{a}'].value)
            for i in range(sheet.nrows):
                if sheet.cell_value(i,8) == int(numAfiliacion):
                    restaurante = str(sheet.cell_value(i,1))
                    break
            break
    textInfoInventario.configure(state=NORMAL)
    textInfoInventario.delete(1.0,END)
    textInfoInventario.insert(END, "                      Inventario")
    textInfoInventario.insert(END, "\n======================================================")
    textInfoInventario.insert(END, "\n\n")
    if(restaurante == ''):
        textInfoInventario.insert(END, f"N/S: {numberSerial} \nNo se encuentra asignado a un restaurante ¯\_(ツ)_/¯")
        textInfoInventario.insert(END, "\n")
    elif(numAfiliacion == ''):
        textInfoInventario.insert(END, f"N/S: {numberSerial} \nSe encuentra en inventario pero no posee afiliacion\npara el restaurante asignado ¯\_(ツ)_/¯")
    else:
        textInfoInventario.insert(END, f"S/N: {numberSerial} corresponde al restaurante: \n{restaurante}")
        textInfoInventario.insert(END, "\n")
    textInfoInventario.insert(END, "\n======================================================")
    textInfoInventario.configure(state=DISABLED)

def cantidadPinpadRest():
    verificarRutaXML(1)
    verificarRutaXML(4)
    cantidad = 0
    restaurante = ''
    try:
        for i in range(sheet.nrows):
            aux = separaCadenaPorCaracter(sheet.cell_value(i,0),'.',0)
            if aux == cajaInventario.get().upper():
                restaurante = str(sheet.cell_value(i,1))
                numAfiliacion = int(sheet.cell_value(i,8))
                for a in range (1,3000):
                    if(hojaInventarioExcel[f'E{a}'].value == numAfiliacion) and (str(hojaInventarioExcel[f'B{a}'].value) == "IPP320 USB" or str(hojaInventarioExcel[f'B{a}'].value) == "VX820" or str(hojaInventarioExcel[f'B{a}'].value) == "IPP320 SERIAL"):#Verifica que sea el restaurante correcto y tiene en cuenta los modelos IPP320 y VX820
                        cantidad += 1
                break
        textInfoInventario.configure(state=NORMAL)
        textInfoInventario.delete(1.0,END)
        textInfoInventario.insert(END, "                      Inventario")
        textInfoInventario.insert(END, "\n======================================================")
        textInfoInventario.insert(END, "\n\n")
        if(restaurante==''):
            textInfoInventario.insert(END, 'El restaurante no existe ¯\_(ツ)_/¯')
        else:
            textInfoInventario.insert(END, f'El restaurate {restaurante} \nPosee {cantidad} pinpad asignados')
        textInfoInventario.insert(END, "\n")
        textInfoInventario.insert(END, "\n======================================================")
        textInfoInventario.configure(state=DISABLED)    
    except:
        textInfoInventario.configure(state=NORMAL)
        textInfoInventario.delete(1.0,END)
        textInfoInventario.insert(END, "                      Inventario")
        textInfoInventario.insert(END, "\n======================================================")
        textInfoInventario.insert(END, "\n\n")
        textInfoInventario.insert(END, "El restaurante no posee afiliación ¯\_(ツ)_/¯")
        textInfoInventario.insert(END, "\n")
        textInfoInventario.insert(END, "\n======================================================")
        textInfoInventario.configure(state=DISABLED)

def buscarLlaves():
    if(cajaIP.get() != ''):

        ip = str(cajaIP.get())
        verificarRutaXML(5)
        for i in range(hojaLlavesExcel.nrows):
            if str(hojaLlavesExcel.cell_value(i,1)) == str(ip):
                llave = str(hojaLlavesExcel.cell_value(i,2))
                terminal = str(hojaLlavesExcel.cell_value(i,0))
                negocio = str(hojaLlavesExcel.cell_value(i,3))
                site = str(separaCadenaPorCaracter(hojaLlavesExcel.cell_value(i,4),'.',0))
                textIP.configure(state=NORMAL)
                textIP.delete(1.0,END)
                textIP.insert(END, "                      Llaves                     ")
                textIP.insert(END, "\n==================================================")
                textIP.insert(END, f"\nSite: {site}\n")
                textIP.insert(END, f'\nLlave: {llave}')
                textIP.insert(END, f"\nNegocio: {negocio}")
                textIP.insert(END, f"\nTerminal: {terminal}")
                textIP.insert(END, "\n\n==================================================")
                textIP.configure(state=DISABLED)
                return
        textIP.configure(state=NORMAL)
        textIP.delete(1.0,END)
        textIP.insert(END, "                      Llaves                     ")
        textIP.insert(END, "\n==================================================")
        textIP.insert(END, "\n\n")
        textIP.insert(END, f'No se encuentran llaves para la ip: {ip}\n')
        textIP.insert(END, "\n\n                    ¯\_(ツ)_/¯")
        textIP.insert(END, "\n")
        textIP.insert(END, "\n==================================================")
        textIP.configure(state=DISABLED)
    else:
        textIP.configure(state=NORMAL)
        textIP.delete(1.0,END)
        textIP.insert(END, "                      Llaves                     ")
        textIP.insert(END, "\n==================================================")
        textIP.insert(END, "\n\n")
        textIP.insert(END, 'Ingrese la información\n')
        textIP.insert(END, "\n\n                    ¯\_(ツ)_/¯")
        textIP.insert(END, "\n")
        textIP.insert(END, "\n==================================================")
        textIP.configure(state=DISABLED)


#Ventana princiapl
ventana = Tk()
ventana.title("Pinpad Control")
ventana.resizable(0,0)
ventana.geometry('+643+333')
pestanias = Notebook(ventana)
pestanias.pack(fill="both",expand="yes")

#Pestania 1
usoNormal=Frame(pestanias)
usoNormal.pack()
pestanias.add(usoNormal,text="Crear Layout")

#Pestania 2
consultaInventario = Frame(pestanias)
consultaInventario.pack()
pestanias.add(consultaInventario,text="Consulta Inventario")

#Pestania 3
llavesPinpad = Frame(pestanias)
llavesPinpad.pack()
pestanias.add(llavesPinpad,text="Llaves")

#Pestania 4
configProgram=Frame(pestanias)
configProgram.pack()
pestanias.add(configProgram,text="Config")

#Pestania creacion Layout
#Entry
etiquetaRestaurate = Label(usoNormal,text="Restaurante: ")
etiquetaRestaurate.grid(row=0,column=1,padx=10,pady=10,sticky="w")
cajaRestaurante = Entry(usoNormal)
cajaRestaurante.grid(row=0,column=2,padx=10,pady=10)

etiquetaTicket = Label(usoNormal,text="Ticket: ")
etiquetaTicket.grid(row=1,column=1,padx=10,pady=10,sticky="w")
cajaTicket = Entry(usoNormal)
cajaTicket.grid(row=1,column=2,padx=10,pady=10)

etiquetaPos = Label(usoNormal,text="Pos: ")
etiquetaPos.grid(row=2,column=1,padx=10,pady=10,sticky="w")
cajaPos = Entry(usoNormal)
cajaPos.grid(row=2,column=2,padx=10,pady=10)

etiquetaSerie = Label(usoNormal,text="Numero de serie de pinpad: ")
etiquetaSerie.grid(row=3,column=1,padx=10,pady=10,sticky="w")
cajaSerie = Entry(usoNormal)
cajaSerie.grid(row=3,column=2,padx=10,pady=10)

etiquetaMotivo = Label(usoNormal,text="Motivo de solicitud: ")
etiquetaMotivo.grid(row=4,column=1,padx=10,pady=10,sticky="w")
cajaMotivo = Entry(usoNormal)
cajaMotivo.grid(row=4,column=2,padx=10,pady=10)

#Opcion exclusiva
modelo = StringVar()
etiquetaModelo = Label(usoNormal,text="Modelo de pinpad: ")
etiquetaModelo.grid(row=0,column=3,padx=10,pady=10)

Radiobutton(usoNormal,text='Ingenico IPP320',variable=modelo, value='Ingenico IPP320').grid(row=0,column=4,sticky="w",padx=10)
Radiobutton(usoNormal,text='Verifone VX820',variable=modelo, value='Verifone VX820').grid(row=1,column=4,sticky="w",padx=10)

inventario=IntVar()
etiquetainv = Label(usoNormal,text="Inventario: ")
etiquetainv.grid(row=2,column=3,padx=10,pady=10,sticky="w")
Radiobutton(usoNormal,text="Pertenece a EvoPayments",variable=inventario,value=2).grid(row=2,column=4,sticky="w",padx=10)
Radiobutton(usoNormal,text="Pertenece a GA",variable=inventario,value=1).grid(row=3,column=4,sticky="w",padx=10)

altaReemplazo=IntVar()
etiquetainv = Label(usoNormal,text="Tipo de solicitud: ")
etiquetainv.grid(row=4,column=3,padx=10,pady=10,sticky="w")
Radiobutton(usoNormal,text="Alta de pinpad",variable=altaReemplazo,value=1).grid(row=4,column=4,sticky="w",padx=10)
Radiobutton(usoNormal,text="Reemplazo de pinpad",variable=altaReemplazo,value=2).grid(row=5,column=4,sticky="w",padx=10)

Button(usoNormal,text="       ✅       ",foreground='green',command=cargarReemplazo).grid(row=7,column=2,padx=10,pady=10,sticky="s")

#Pestaña de Consulta de inventario
etiquetaInvetario=Label(consultaInventario,text="").grid(row=0,column=0,sticky="n,w",padx=10,pady=10)
cajaInventario=Entry(consultaInventario)
cajaInventario.grid(row=0,column=1,sticky="w,n",padx=10,pady=10)
botonSerial = Button(consultaInventario,text="Serial Number",command=buscarPinpadInventario).grid(row=0,column=1,padx=10)
textInfoInventario = Text(consultaInventario,width=54,height=7,state=DISABLED)
textInfoInventario.grid(row=0,column=2,sticky="w",padx=10,pady=10)
botonCantidadInventario= Button(consultaInventario,text="Cantidad",command=cantidadPinpadRest).grid(row=0,column=1,padx=10,sticky="s")

#Pestania de Llaves
Label(llavesPinpad,text="IP: ").grid(row=0,column=0,sticky="n,w",padx=10,pady=10)
cajaIP = Entry(llavesPinpad)
cajaIP.grid(row=0,column=1,sticky="w,n",padx=10,pady=10)
botonIP= Button(llavesPinpad,text="Buscar",command=buscarLlaves).grid(row=0,column=1,padx=10)
textIP = Text(llavesPinpad,width=50,height=9,state=DISABLED)
textIP.grid(row=0,column=2,sticky="w",padx=5,pady=10)

##Pestaña de configProgram##
setCajaInfoRest=StringVar()
etiquetaInfoRest=Label(configProgram,text="Información de Restaurantes: ").grid(row=0,column=1,sticky="w",padx=10,pady=10)
cajaInfoRest=Entry(configProgram,textvariable=setCajaInfoRest,state=DISABLED).grid(row=0,column=2,sticky="w",padx=10,pady=10)
botonInfoRest = Button(configProgram,text="⚙️",command=lambda: verificarRutaXML(1,True)).grid(row=0,column=3,sticky="w")

setCajaPlantilla=StringVar()
etiquetaPlantilla=Label(configProgram,text="Plantilla: ").grid(row=1,column=1,sticky="w",padx=10,pady=10)
cajaPlantilla=Entry(configProgram,textvariable=setCajaPlantilla,state=DISABLED).grid(row=1,column=2,sticky="w",padx=10,pady=10)
botonPlantilla = Button(configProgram,text="⚙️",command= lambda: verificarRutaXML(2,True)).grid(row=1,column=3,sticky="w")

setCajaGuardar=StringVar()
etiquetaGuardar=Label(configProgram,text="Guardar excel como: ").grid(row=2,column=1,sticky="w",padx=10,pady=10)
cajaGuardar=Entry(configProgram,textvariable=setCajaGuardar,state=DISABLED).grid(row=2,column=2,sticky="w",padx=10,pady=10)
botonGuardar = Button(configProgram,text="⚙️",command=lambda: verificarRutaXML(3,True)).grid(row=2,column=3,sticky="w")

setCajaInventario=StringVar()
etiquetaInventario=Label(configProgram,text="Inventario de pinpad: ").grid(row=3,column=1,sticky="w",padx=10,pady=10)
cajaRutaInventario=Entry(configProgram,textvariable=setCajaInventario,state=DISABLED).grid(row=3,column=2,sticky="w",padx=10,pady=10)
botonInventario = Button(configProgram,text="⚙️",command=lambda: verificarRutaXML(4,True)).grid(row=3,column=3,sticky="w")

setCajaLlaves=StringVar()
etiquetaLlaves=Label(configProgram,text="Información de llaves: ").grid(row=4,column=1,sticky="w",padx=10,pady=10)
cajaLlaves= Entry(configProgram,textvariable=setCajaLlaves,state=DISABLED).grid(row=4,column=2,sticky="w",padx=10,pady=10)
botonLlaves = Button(configProgram,text="⚙️",command=lambda: verificarRutaXML(5,True)).grid(row=4,column=3,sticky="w")

Button(configProgram,text="❔",command=copirait).place(x=595,y=240)

ventana.mainloop()
